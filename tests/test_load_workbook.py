import openpyxl

from import_workbook import load_workbook


def _write_sheet(wb, name, header, rows):
    ws = wb.create_sheet(name)
    ws.append(header)
    for row in rows:
        ws.append(row)


def build_sample_workbook(path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    _write_sheet(
        wb, "Customers",
        ["customer_name", "company_name", "currency", "email", "phone", "bill_address_line1", "memo"],
        [["Acme Corp", "Acme Corp Pte Ltd", "CHF", "ap@acme.test", "555-1000", "1 Main St", "VIP"]],
    )
    _write_sheet(
        wb, "Vendors",
        ["vendor_name", "currency", "address_line1"],
        [["Acme Supplier", "USD", "9 Vendor Rd"]],
    )
    _write_sheet(
        wb, "Invoices",
        ["invoice_number", "customer_name", "invoice_date", "currency", "exchange_rate", "item_name", "quantity", "rate", "amount"],
        [
            ["INV-1001", "Acme Corp", "2026-07-01", "CHF", 1.0962, "Consulting", 10, 150, 1500],
            ["INV-1001", "Acme Corp", "2026-07-01", "CHF", 1.0962, "Product A", "", "", 200],
        ],
    )
    _write_sheet(
        wb, "Bills",
        ["bill_number", "vendor_name", "bill_date", "account", "amount"],
        [["B-100", "Acme Supplier", "2026-07-02", "Office Supplies", 250]],
    )
    _write_sheet(
        wb, "CustomerPayments",
        ["payment_id", "customer_name", "payment_date", "deposit_to_account", "applied_invoice_number", "applied_amount"],
        [["p1", "Acme Corp", "2026-07-15", "Undeposited Funds", "INV-1001", 1700]],
    )
    _write_sheet(
        wb, "VendorPayments",
        ["payment_id", "vendor_name", "payment_date", "bank_account", "applied_bill_number", "applied_amount"],
        [["vp1", "Acme Supplier", "2026-07-15", "Bank Account", "B-100", 250]],
    )

    wb.save(path)


def test_load_workbook_round_trips_all_six_sheets(tmp_path):
    path = tmp_path / "sample.xlsx"
    build_sample_workbook(path)

    wb = load_workbook(path)

    assert len(wb.customers) == 1
    assert wb.customers[0].name == "Acme Corp"
    assert wb.customers[0].address_lines == ["1 Main St"]

    assert len(wb.vendors) == 1
    assert wb.vendors[0].name == "Acme Supplier"

    assert len(wb.invoices) == 1
    invoice = wb.invoices[0]
    assert invoice.ref_number == "INV-1001"
    assert invoice.currency == "CHF"
    assert len(invoice.lines) == 2
    assert invoice.lines[0].item == "Consulting"
    assert invoice.lines[0].amount == 1500

    assert len(wb.bills) == 1
    assert wb.bills[0].ref_number == "B-100"
    assert wb.bills[0].lines[0].account == "Office Supplies"

    assert len(wb.customer_payments) == 1
    assert wb.customer_payments[0].applications[0].invoice_ref == "INV-1001"

    assert len(wb.vendor_payments) == 1
    assert wb.vendor_payments[0].applications[0].bill_ref == "B-100"


def test_load_workbook_tolerates_missing_sheets(tmp_path):
    path = tmp_path / "customers_only.xlsx"
    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)
    _write_sheet(wb_out, "Customers", ["customer_name"], [["Solo Customer"]])
    wb_out.save(path)

    wb = load_workbook(path)
    assert len(wb.customers) == 1
    assert wb.vendors == []
    assert wb.invoices == []
    assert wb.customer_payments == []


def test_load_workbook_missing_file_raises():
    import pytest

    with pytest.raises(FileNotFoundError):
        load_workbook(__import__("pathlib").Path("/nonexistent/workbook.xlsx"))
