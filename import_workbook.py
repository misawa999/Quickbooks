#!/usr/bin/env python3
"""Import an Excel workbook of business transactions into QuickBooks Desktop.

Sibling to import_batch.py (which handles the JSON journal-entry batch
format) -- this handles a separate .xlsx workbook format covering
Customers, Vendors, Invoices, Bills, CustomerPayments and VendorPayments,
each on its own optional sheet. Both tools share the same dedupe log
(dedupe.py, import_log.jsonl) via namespaced line_ids (cust-/vend-/inv-/
bill-/pmt-/vpmt- vs. journal entries' unprefixed line_id) so they can be run
side by side without colliding.

Usage:
    python import_workbook.py workbook.xlsx                # dry run (default)
    python import_workbook.py workbook.xlsx --commit        # write to QuickBooks

Processing order matters: Customers and Vendors are created before
Invoices/Bills that reference them, which are created before
CustomerPayments/VendorPayments that settle them. Payments resolve their
target invoice/bill RefNumbers to QuickBooks TxnIDs via a live query
(txn_lookup.py) at commit time -- see that module for why this can't be
done from the workbook data alone.
"""
from __future__ import annotations

import argparse
import sys
from collections import OrderedDict
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set, Tuple

import openpyxl
from pydantic import ValidationError

from dedupe import load_processed_line_ids
from import_batch import DEFAULT_LOG_PATH, log_safely
from preflight import (
    existing_customer_names,
    existing_vendor_names,
    missing_accounts,
    missing_items,
)
from qb_requests import (
    build_bill_add_rq,
    build_bill_payment_check_add_rq,
    build_customer_add_rq,
    build_invoice_add_rq,
    build_receive_payment_add_rq,
    build_vendor_add_rq,
    parse_bill_add_rs,
    parse_bill_payment_check_add_rs,
    parse_customer_add_rs,
    parse_invoice_add_rs,
    parse_receive_payment_add_rs,
    parse_vendor_add_rs,
)
from qb_session import QBSession, QBSessionError
from schema import (
    Bill,
    Customer,
    CustomerPayment,
    Invoice,
    Vendor,
    VendorPayment,
    WorkbookBatch,
)
from txn_lookup import TxnResolutionError, resolve_bill_txn_id, resolve_invoice_txn_id


# -- Excel loading ----------------------------------------------------------

def _read_sheet(wb, sheet_name: str) -> List[dict]:
    """Read a sheet into a list of {header: value} dicts. Missing sheets
    yield an empty list -- every sheet in this workbook format is optional."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    records: List[dict] = []
    for row in rows_iter:
        if all(v is None for v in row):
            continue
        record = {headers[i]: row[i] for i in range(len(headers)) if headers[i]}
        records.append(record)
    return records


def _to_str(value) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    text = str(value).strip()
    return text or None


def _to_decimal(value) -> Decimal:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        raise ValueError("missing required amount")
    return Decimal(str(value))


def _to_decimal_or_none(value) -> Optional[Decimal]:
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    return Decimal(str(value))


def _address_lines(row: dict, prefix: str) -> List[str]:
    lines = []
    for i in range(1, 6):
        value = _to_str(row.get(f"{prefix}{i}"))
        if value:
            lines.append(value)
    return lines


def _normalize(value):
    if value is None:
        return ""
    if isinstance(value, str):
        return value.strip()
    return value


def group_tidy_rows(
    rows: List[dict], key_col: str, header_cols: List[str], line_cols: List[str]
) -> List[dict]:
    """Group tidy/long rows (one row per line item) into
    {"key": ..., "header": {...}, "lines": [{...}, ...]} groups, keyed by
    key_col. Every row sharing a key must agree on every header_col --
    raises a clear error naming the key and the conflicting column if not."""
    groups: "OrderedDict[str, dict]" = OrderedDict()
    for row in rows:
        raw_key = row.get(key_col)
        if raw_key is None or str(raw_key).strip() == "":
            raise ValueError(f"row is missing required {key_col!r}: {row}")
        key = str(raw_key).strip()
        if key not in groups:
            groups[key] = {
                "key": key,
                "header": {col: row.get(col) for col in header_cols},
                "lines": [],
            }
        else:
            existing_header = groups[key]["header"]
            for col in header_cols:
                if _normalize(existing_header.get(col)) != _normalize(row.get(col)):
                    raise ValueError(
                        f"{key_col}={key!r}: inconsistent {col!r} across rows "
                        f"({existing_header.get(col)!r} vs {row.get(col)!r})"
                    )
        groups[key]["lines"].append({col: row.get(col) for col in line_cols})
    return list(groups.values())


def _row_to_customer(row: dict) -> dict:
    return {
        "name": _to_str(row.get("customer_name")),
        "company_name": _to_str(row.get("company_name")),
        "currency": _to_str(row.get("currency")),
        "email": _to_str(row.get("email")),
        "phone": _to_str(row.get("phone")),
        "address_lines": _address_lines(row, "bill_address_line"),
        "memo": _to_str(row.get("memo")),
    }


def _row_to_vendor(row: dict) -> dict:
    return {
        "name": _to_str(row.get("vendor_name")),
        "company_name": _to_str(row.get("company_name")),
        "currency": _to_str(row.get("currency")),
        "email": _to_str(row.get("email")),
        "phone": _to_str(row.get("phone")),
        "address_lines": _address_lines(row, "address_line"),
        "memo": _to_str(row.get("memo")),
    }


def _group_to_invoice(group: dict) -> dict:
    header = group["header"]
    lines = [
        {
            "item": _to_str(line.get("item_name")),
            "description": _to_str(line.get("description")),
            "quantity": _to_decimal_or_none(line.get("quantity")),
            "rate": _to_decimal_or_none(line.get("rate")),
            "amount": _to_decimal(line.get("amount")),
        }
        for line in group["lines"]
    ]
    return {
        "ref_number": _to_str(group["key"]),
        "customer": _to_str(header.get("customer_name")),
        "date": header.get("invoice_date"),
        "due_date": header.get("due_date") or None,
        "terms": _to_str(header.get("terms")),
        "currency": _to_str(header.get("currency")),
        "exchange_rate": _to_decimal_or_none(header.get("exchange_rate")),
        "ar_account": _to_str(header.get("ar_account")),
        "memo": _to_str(header.get("memo")),
        "lines": lines,
    }


def _group_to_bill(group: dict) -> dict:
    header = group["header"]
    lines = [
        {
            "account": _to_str(line.get("account")),
            "description": _to_str(line.get("description")),
            "amount": _to_decimal(line.get("amount")),
        }
        for line in group["lines"]
    ]
    return {
        "ref_number": _to_str(group["key"]),
        "vendor": _to_str(header.get("vendor_name")),
        "date": header.get("bill_date"),
        "due_date": header.get("due_date") or None,
        "currency": _to_str(header.get("currency")),
        "exchange_rate": _to_decimal_or_none(header.get("exchange_rate")),
        "ap_account": _to_str(header.get("ap_account")),
        "memo": _to_str(header.get("memo")),
        "lines": lines,
    }


def _group_to_customer_payment(group: dict) -> dict:
    header = group["header"]
    applications = [
        {
            "invoice_ref": _to_str(line.get("applied_invoice_number")),
            "amount": _to_decimal(line.get("applied_amount")),
        }
        for line in group["lines"]
    ]
    return {
        "payment_id": _to_str(group["key"]),
        "customer": _to_str(header.get("customer_name")),
        "date": header.get("payment_date"),
        "deposit_to_account": _to_str(header.get("deposit_to_account")),
        "ar_account": _to_str(header.get("ar_account")),
        "currency": _to_str(header.get("currency")),
        "exchange_rate": _to_decimal_or_none(header.get("exchange_rate")),
        "payment_method": _to_str(header.get("payment_method")),
        "ref_number": _to_str(header.get("ref_number")),
        "memo": _to_str(header.get("memo")),
        "applications": applications,
    }


def _group_to_vendor_payment(group: dict) -> dict:
    header = group["header"]
    applications = [
        {
            "bill_ref": _to_str(line.get("applied_bill_number")),
            "amount": _to_decimal(line.get("applied_amount")),
        }
        for line in group["lines"]
    ]
    return {
        "payment_id": _to_str(group["key"]),
        "vendor": _to_str(header.get("vendor_name")),
        "date": header.get("payment_date"),
        "bank_account": _to_str(header.get("bank_account")),
        "ap_account": _to_str(header.get("ap_account")),
        "currency": _to_str(header.get("currency")),
        "exchange_rate": _to_decimal_or_none(header.get("exchange_rate")),
        "check_number": _to_str(header.get("check_number")),
        "memo": _to_str(header.get("memo")),
        "applications": applications,
    }


def load_workbook(path: Path) -> WorkbookBatch:
    if not path.exists():
        raise FileNotFoundError(f"workbook not found: {path}")
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        customers = [Customer.model_validate(_row_to_customer(r)) for r in _read_sheet(wb, "Customers")]
        vendors = [Vendor.model_validate(_row_to_vendor(r)) for r in _read_sheet(wb, "Vendors")]

        invoice_groups = group_tidy_rows(
            _read_sheet(wb, "Invoices"),
            "invoice_number",
            header_cols=[
                "customer_name", "invoice_date", "due_date", "terms",
                "currency", "exchange_rate", "ar_account", "memo",
            ],
            line_cols=["item_name", "description", "quantity", "rate", "amount"],
        )
        invoices = [Invoice.model_validate(_group_to_invoice(g)) for g in invoice_groups]

        bill_groups = group_tidy_rows(
            _read_sheet(wb, "Bills"),
            "bill_number",
            header_cols=["vendor_name", "bill_date", "due_date", "currency", "exchange_rate", "ap_account", "memo"],
            line_cols=["account", "description", "amount"],
        )
        bills = [Bill.model_validate(_group_to_bill(g)) for g in bill_groups]

        cust_pmt_groups = group_tidy_rows(
            _read_sheet(wb, "CustomerPayments"),
            "payment_id",
            header_cols=[
                "customer_name", "payment_date", "deposit_to_account", "ar_account",
                "currency", "exchange_rate", "payment_method", "ref_number", "memo",
            ],
            line_cols=["applied_invoice_number", "applied_amount"],
        )
        customer_payments = [
            CustomerPayment.model_validate(_group_to_customer_payment(g)) for g in cust_pmt_groups
        ]

        vend_pmt_groups = group_tidy_rows(
            _read_sheet(wb, "VendorPayments"),
            "payment_id",
            header_cols=[
                "vendor_name", "payment_date", "bank_account", "ap_account",
                "currency", "exchange_rate", "check_number", "memo",
            ],
            line_cols=["applied_bill_number", "applied_amount"],
        )
        vendor_payments = [
            VendorPayment.model_validate(_group_to_vendor_payment(g)) for g in vend_pmt_groups
        ]
    finally:
        wb.close()

    return WorkbookBatch(
        customers=customers,
        vendors=vendors,
        invoices=invoices,
        bills=bills,
        customer_payments=customer_payments,
        vendor_payments=vendor_payments,
    )


# -- Preflight helpers --------------------------------------------------

def all_account_names(wb: WorkbookBatch) -> List[str]:
    names: List[str] = []
    for bill in wb.bills:
        names.extend(line.account for line in bill.lines)
        if bill.ap_account:
            names.append(bill.ap_account)
    for inv in wb.invoices:
        if inv.ar_account:
            names.append(inv.ar_account)
    for p in wb.customer_payments:
        names.append(p.deposit_to_account)
        if p.ar_account:
            names.append(p.ar_account)
    for p in wb.vendor_payments:
        names.append(p.bank_account)
        if p.ap_account:
            names.append(p.ap_account)
    return names


def all_item_names(wb: WorkbookBatch) -> List[str]:
    names: List[str] = []
    for inv in wb.invoices:
        names.extend(line.item for line in inv.lines)
    return names


def preview_invoice_resolutions(
    session: QBSession, customer_payments: List[CustomerPayment]
) -> Dict[Tuple[str, str], str]:
    results: Dict[Tuple[str, str], str] = {}
    for p in customer_payments:
        for app in p.applications:
            key = (p.customer, app.invoice_ref)
            if key in results:
                continue
            try:
                results[key] = resolve_invoice_txn_id(session, app.invoice_ref, p.customer)
            except TxnResolutionError as e:
                results[key] = f"ERROR: {e}"
    return results


def preview_bill_resolutions(
    session: QBSession, vendor_payments: List[VendorPayment]
) -> Dict[Tuple[str, str], str]:
    results: Dict[Tuple[str, str], str] = {}
    for p in vendor_payments:
        for app in p.applications:
            key = (p.vendor, app.bill_ref)
            if key in results:
                continue
            try:
                results[key] = resolve_bill_txn_id(session, app.bill_ref, p.vendor)
            except TxnResolutionError as e:
                results[key] = f"ERROR: {e}"
    return results


# -- Dry-run report -------------------------------------------------------

def print_workbook_dry_run_report(
    wb: WorkbookBatch,
    skip_ids: Set[str],
    missing_account_names: List[str],
    missing_item_names: List[str],
    existing_customers: Set[str],
    existing_vendors: Set[str],
    invoice_resolution: Dict[Tuple[str, str], str],
    bill_resolution: Dict[Tuple[str, str], str],
    emit: Callable[[str], None] = print,
) -> None:
    if missing_account_names:
        emit("PREFLIGHT FAILED - accounts not found in QuickBooks:")
        for name in missing_account_names:
            emit(f"  - {name}")
        emit("")
    if missing_item_names:
        emit("PREFLIGHT FAILED - items not found in QuickBooks:")
        for name in missing_item_names:
            emit(f"  - {name}")
        emit("")

    emit(f"Customers ({len(wb.customers)})")
    for c in wb.customers:
        tag = ""
        if c.name in existing_customers:
            tag = "  [ALREADY EXISTS IN QB - will skip]"
        elif f"cust-{c.name}" in skip_ids:
            tag = "  [DUPLICATE - already imported]"
        cur = f"  currency={c.currency}" if c.currency else ""
        emit(f"  {c.name}{cur}{tag}")
    emit("")

    emit(f"Vendors ({len(wb.vendors)})")
    for v in wb.vendors:
        tag = ""
        if v.name in existing_vendors:
            tag = "  [ALREADY EXISTS IN QB - will skip]"
        elif f"vend-{v.name}" in skip_ids:
            tag = "  [DUPLICATE - already imported]"
        cur = f"  currency={v.currency}" if v.currency else ""
        emit(f"  {v.name}{cur}{tag}")
    emit("")

    emit(f"Invoices ({len(wb.invoices)})")
    for inv in wb.invoices:
        dup = "  [DUPLICATE - already imported]" if f"inv-{inv.customer}-{inv.ref_number}" in skip_ids else ""
        cur = f"{inv.currency} @ {inv.exchange_rate}" if inv.currency else "home currency"
        emit(f"  [{inv.ref_number}] {inv.customer}  {inv.date}  {cur}{dup}")
        for line in inv.lines:
            emit(f"      {line.amount:.2f}  {line.item}")
    emit("")

    emit(f"Bills ({len(wb.bills)})")
    for bill in wb.bills:
        dup = "  [DUPLICATE - already imported]" if f"bill-{bill.vendor}-{bill.ref_number}" in skip_ids else ""
        cur = f"{bill.currency} @ {bill.exchange_rate}" if bill.currency else "home currency"
        emit(f"  [{bill.ref_number}] {bill.vendor}  {bill.date}  {cur}{dup}")
        for line in bill.lines:
            emit(f"      {line.amount:.2f}  {line.account}")
    emit("")

    emit(f"CustomerPayments ({len(wb.customer_payments)})")
    for p in wb.customer_payments:
        dup = "  [DUPLICATE - already imported]" if f"pmt-{p.payment_id}" in skip_ids else ""
        cur = f"{p.currency} @ {p.exchange_rate}" if p.currency else "home currency"
        emit(f"  [{p.payment_id}] {p.customer}  {p.date}  {cur}{dup}")
        for app in p.applications:
            resolved = invoice_resolution.get((p.customer, app.invoice_ref))
            if resolved is None:
                status = "[not checked - QuickBooks unreachable]"
            elif resolved.startswith("ERROR:"):
                status = f"[{resolved}]"
            else:
                status = f"[-> TxnID {resolved}]"
            emit(f"      {app.amount:.2f}  applied to invoice {app.invoice_ref} {status}")
    emit("")

    emit(f"VendorPayments ({len(wb.vendor_payments)})")
    for p in wb.vendor_payments:
        dup = "  [DUPLICATE - already imported]" if f"vpmt-{p.payment_id}" in skip_ids else ""
        cur = f"{p.currency} @ {p.exchange_rate}" if p.currency else "home currency"
        emit(f"  [{p.payment_id}] {p.vendor}  {p.date}  {cur}{dup}")
        for app in p.applications:
            resolved = bill_resolution.get((p.vendor, app.bill_ref))
            if resolved is None:
                status = "[not checked - QuickBooks unreachable]"
            elif resolved.startswith("ERROR:"):
                status = f"[{resolved}]"
            else:
                status = f"[-> TxnID {resolved}]"
            emit(f"      {app.amount:.2f}  applied to bill {app.bill_ref} {status}")
    emit("")


# -- Commit -----------------------------------------------------------------

_ENTITY_TYPES = ["customer", "vendor", "invoice", "bill", "customer_payment", "vendor_payment"]


def run_workbook_commit(
    wb: WorkbookBatch,
    company_file: str,
    log_path: Path,
    skip_ids: Set[str],
    force: bool,
    continue_on_error: bool,
    emit: Callable[[str], None] = print,
) -> int:
    tallies: Dict[str, List[int]] = {t: [0, 0, 0] for t in _ENTITY_TYPES}  # [inserted, skipped, failed]

    def finish() -> int:
        emit("")
        total_failed = 0
        for entity_type in _ENTITY_TYPES:
            ins, skip, fail = tallies[entity_type]
            emit(f"{entity_type}: inserted={ins} skipped={skip} failed={fail}")
            total_failed += fail
        return 0 if total_failed == 0 else 1

    def log(entity_type: str, key: str, status: str, txn_id, status_code, message: str) -> None:
        log_safely(
            log_path,
            {
                "ts": datetime.now(timezone.utc).isoformat(),
                "entity_type": entity_type,
                "line_id": key,
                "status": status,
                "qb_txn_id": txn_id,
                "qbxml_status_code": status_code,
                "message": message,
            },
            emit=emit,
        )

    def fail_one(key: str, entity_type: str, message: str) -> None:
        tallies[entity_type][2] += 1
        emit(f"[{key}] ERROR: {message}")
        log(entity_type, key, "error", None, None, message)

    def commit_one(key: str, entity_type: str, send: Callable[[], dict]) -> bool:
        """Runs `send()` (build request, process, parse response), logs the
        outcome, and returns True if the whole batch should stop (a hard
        error or a QuickBooks-rejected FAILED status)."""
        if key in skip_ids and not force:
            emit(f"[{key}] skipped (already imported)")
            tallies[entity_type][1] += 1
            return False
        try:
            result = send()
        except Exception as e:
            fail_one(key, entity_type, str(e))
            return True
        if result["status_code"] == 0:
            tallies[entity_type][0] += 1
            ref_id = result.get("txn_id") or result.get("list_id")
            emit(f"[{key}] OK" + (f" -> {ref_id}" if ref_id else ""))
            log(entity_type, key, "ok", ref_id, result["status_code"], result["status_message"])
            return False
        tallies[entity_type][2] += 1
        emit(f"[{key}] FAILED ({result['status_code']}): {result['status_message']}")
        log(entity_type, key, "error", None, result["status_code"], result["status_message"])
        return True

    with QBSession(company_file) as session:
        missing_acc = missing_accounts(session, all_account_names(wb))
        missing_it = missing_items(session, all_item_names(wb))
        if missing_acc or missing_it:
            emit("Aborting: preflight failed.")
            for name in missing_acc:
                emit(f"  account not found: {name}")
            for name in missing_it:
                emit(f"  item not found: {name}")
            return 1

        existing_customers = existing_customer_names(session, [c.name for c in wb.customers])
        for c in wb.customers:
            key = f"cust-{c.name}"
            if c.name in existing_customers:
                emit(f"[{key}] skipped (already exists in QuickBooks)")
                tallies["customer"][1] += 1
                continue
            stop = commit_one(
                key, "customer",
                lambda c=c: parse_customer_add_rs(session.process(build_customer_add_rq(c))),
            )
            if stop and not continue_on_error:
                return finish()

        existing_vendors = existing_vendor_names(session, [v.name for v in wb.vendors])
        for v in wb.vendors:
            key = f"vend-{v.name}"
            if v.name in existing_vendors:
                emit(f"[{key}] skipped (already exists in QuickBooks)")
                tallies["vendor"][1] += 1
                continue
            stop = commit_one(
                key, "vendor",
                lambda v=v: parse_vendor_add_rs(session.process(build_vendor_add_rq(v))),
            )
            if stop and not continue_on_error:
                return finish()

        for inv in wb.invoices:
            key = f"inv-{inv.customer}-{inv.ref_number}"
            stop = commit_one(
                key, "invoice",
                lambda inv=inv: parse_invoice_add_rs(session.process(build_invoice_add_rq(inv))),
            )
            if stop and not continue_on_error:
                return finish()

        for bill in wb.bills:
            key = f"bill-{bill.vendor}-{bill.ref_number}"
            stop = commit_one(
                key, "bill",
                lambda bill=bill: parse_bill_add_rs(session.process(build_bill_add_rq(bill))),
            )
            if stop and not continue_on_error:
                return finish()

        for p in wb.customer_payments:
            key = f"pmt-{p.payment_id}"
            if key in skip_ids and not force:
                emit(f"[{key}] skipped (already imported)")
                tallies["customer_payment"][1] += 1
                continue
            txn_ids: Dict[str, str] = {}
            errors: List[str] = []
            for app in p.applications:
                try:
                    txn_ids[app.invoice_ref] = resolve_invoice_txn_id(session, app.invoice_ref, p.customer)
                except TxnResolutionError as e:
                    errors.append(str(e))
            if errors:
                fail_one(key, "customer_payment", "; ".join(errors))
                if not continue_on_error:
                    return finish()
                continue
            stop = commit_one(
                key, "customer_payment",
                lambda p=p, txn_ids=txn_ids: parse_receive_payment_add_rs(
                    session.process(build_receive_payment_add_rq(p, txn_ids))
                ),
            )
            if stop and not continue_on_error:
                return finish()

        for p in wb.vendor_payments:
            key = f"vpmt-{p.payment_id}"
            if key in skip_ids and not force:
                emit(f"[{key}] skipped (already imported)")
                tallies["vendor_payment"][1] += 1
                continue
            txn_ids = {}
            errors = []
            for app in p.applications:
                try:
                    txn_ids[app.bill_ref] = resolve_bill_txn_id(session, app.bill_ref, p.vendor)
                except TxnResolutionError as e:
                    errors.append(str(e))
            if errors:
                fail_one(key, "vendor_payment", "; ".join(errors))
                if not continue_on_error:
                    return finish()
                continue
            stop = commit_one(
                key, "vendor_payment",
                lambda p=p, txn_ids=txn_ids: parse_bill_payment_check_add_rs(
                    session.process(build_bill_payment_check_add_rq(p, txn_ids))
                ),
            )
            if stop and not continue_on_error:
                return finish()

    return finish()


def main() -> int:
    parser = argparse.ArgumentParser(
        description=(
            "Import an Excel workbook of customers/vendors/invoices/bills/"
            "payments into QuickBooks Desktop."
        )
    )
    parser.add_argument("workbook_file", type=Path)
    parser.add_argument(
        "--company-file",
        default="",
        help="Path to the .qbw company file (blank = currently open file)",
    )
    parser.add_argument(
        "--log-file",
        type=Path,
        default=DEFAULT_LOG_PATH,
        help=f"Where to write the import log (default: {DEFAULT_LOG_PATH})",
    )
    parser.add_argument(
        "--commit", action="store_true", help="Actually write to QuickBooks. Default is dry-run."
    )
    parser.add_argument(
        "--force", action="store_true", help="Re-import records already marked ok in the log."
    )
    parser.add_argument(
        "--continue-on-error",
        action="store_true",
        help="Keep going after a failed insert instead of stopping.",
    )
    args = parser.parse_args()

    try:
        wb = load_workbook(args.workbook_file)
    except ValidationError as e:
        print("Workbook validation failed:")
        print(e)
        return 1
    except (ValueError, FileNotFoundError) as e:
        print(f"Could not read workbook file: {e}")
        return 1

    print(f"Log file: {args.log_file}\n")
    skip_ids = set() if args.force else load_processed_line_ids(args.log_file)

    if not args.commit:
        missing_acc: List[str] = []
        missing_it: List[str] = []
        existing_cust: Set[str] = set()
        existing_vend: Set[str] = set()
        invoice_resolution: Dict[Tuple[str, str], str] = {}
        bill_resolution: Dict[Tuple[str, str], str] = {}
        try:
            with QBSession(args.company_file) as session:
                missing_acc = missing_accounts(session, all_account_names(wb))
                missing_it = missing_items(session, all_item_names(wb))
                existing_cust = existing_customer_names(session, [c.name for c in wb.customers])
                existing_vend = existing_vendor_names(session, [v.name for v in wb.vendors])
                invoice_resolution = preview_invoice_resolutions(session, wb.customer_payments)
                bill_resolution = preview_bill_resolutions(session, wb.vendor_payments)
        except QBSessionError as e:
            print(f"(preflight skipped - could not connect to QuickBooks: {e})\n")
        print_workbook_dry_run_report(
            wb, skip_ids, missing_acc, missing_it, existing_cust, existing_vend,
            invoice_resolution, bill_resolution,
        )
        print("Dry run only - nothing was written. Re-run with --commit to import.")
        return 0

    try:
        return run_workbook_commit(
            wb, args.company_file, args.log_file, skip_ids, args.force, args.continue_on_error
        )
    except QBSessionError as e:
        print(f"Could not connect to QuickBooks: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
